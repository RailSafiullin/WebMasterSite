from typing import Generator

from openpyxl import load_workbook

from api.auth.schemas import UserCreate, UserCreateWithRole

def import_users_from_excel(file) -> Generator[UserCreateWithRole, None, None]:
    """
    Generator, which converts excel row to UserCreate schema
    :param file: is FileObject
    """
    wb = load_workbook(file, read_only=True, data_only=True)
    for row in wb.active.iter_rows(values_only=True):
        email, username, password, role = map(lambda value: value if value is None else str(value), row)
        yield UserCreateWithRole(id=-1, email=email, username=username, password="None", role=role)